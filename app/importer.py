from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .db import db, utcnow
from .phones import normalize_in_mobile

NULLISH = {"", "NULL", "null", "None", "none"}


def clean(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float):
        if value != value:
            return None
        if value.is_integer():
            value = int(value)
    text = str(value).strip()
    if text in NULLISH:
        return None
    return text


def rotary_key(value) -> str | None:
    text = clean(value)
    if not text:
        return None
    return text


def extra_from_row(header: tuple, row: tuple) -> dict:
    skip = {"Name", "Mobile", "Email", "RotaryId", "Club_Name", "Club_No"}
    out = {}
    for key, val in zip(header, row):
        if key in skip:
            continue
        cleaned = clean(val)
        if cleaned is not None:
            out[str(key)] = cleaned
    return out


def _index_existing(conn) -> tuple[dict, dict, dict]:
    by_rotary = {}
    by_mobile = {}
    by_name_email = {}
    rows = conn.execute(
        "SELECT id, club_no, name, mobile_digits, email, rotary_id FROM members"
    ).fetchall()
    for r in rows:
        if r["rotary_id"]:
            by_rotary[r["rotary_id"]] = r["id"]
        if r["club_no"] and r["mobile_digits"]:
            by_mobile[(r["club_no"], r["mobile_digits"])] = r["id"]
        email = (r["email"] or "").lower()
        name = (r["name"] or "").lower()
        if r["club_no"] and name:
            by_name_email[(r["club_no"], name, email)] = r["id"]
    return by_rotary, by_mobile, by_name_email


def import_excel(path: str | Path) -> dict:
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    header_index = {name: i for i, name in enumerate(header) if name}

    required = ["Name", "Club_Name", "Club_No"]
    missing = [c for c in required if c not in header_index]
    if missing:
        raise ValueError(f"Excel is missing columns: {', '.join(missing)}")

    now = utcnow()
    inserted = updated = unchanged = skipped = 0
    clubs_upserted = 0
    seen_clubs: dict[str, str] = {}

    parsed = []
    for row in rows_iter:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def col(name):
            i = header_index.get(name)
            return row[i] if i is not None and i < len(row) else None

        name = clean(col("Name"))
        club_name = clean(col("Club_Name"))
        club_no = clean(col("Club_No"))
        if not name or not club_name or not club_no:
            skipped += 1
            continue
        mobile_raw = col("Mobile")
        ten, _wa = normalize_in_mobile(mobile_raw)
        parsed.append(
            {
                "name": name,
                "club_name": club_name,
                "club_no": club_no,
                "mobile": ten or clean(mobile_raw),
                "mobile_digits": ten,
                "email": clean(col("Email")),
                "rotary_id": rotary_key(col("RotaryId")),
                "extra": extra_from_row(header, row),
            }
        )
        seen_clubs[club_no] = club_name

    with db() as conn:
        for club_no, club_name in seen_clubs.items():
            existing = conn.execute(
                "SELECT club_name FROM clubs WHERE club_no = ?", (club_no,)
            ).fetchone()
            if existing is None:
                conn.execute(
                    "INSERT INTO clubs (club_no, club_name, updated_at) VALUES (?, ?, ?)",
                    (club_no, club_name, now),
                )
                clubs_upserted += 1
            elif existing["club_name"] != club_name:
                conn.execute(
                    "UPDATE clubs SET club_name = ?, updated_at = ? WHERE club_no = ?",
                    (club_name, now, club_no),
                )
                clubs_upserted += 1

        by_rotary, by_mobile, by_name_email = _index_existing(conn)

        for item in parsed:
            extra_json = json.dumps(item["extra"], ensure_ascii=False)
            match_id = None
            if item["rotary_id"] and item["rotary_id"] in by_rotary:
                match_id = by_rotary[item["rotary_id"]]
            elif item["mobile_digits"] and (item["club_no"], item["mobile_digits"]) in by_mobile:
                match_id = by_mobile[(item["club_no"], item["mobile_digits"])]
            elif (item["club_no"], item["name"].lower(), (item["email"] or "").lower()) in by_name_email:
                match_id = by_name_email[(item["club_no"], item["name"].lower(), (item["email"] or "").lower())]

            if match_id is None:
                conn.execute(
                    """
                    INSERT INTO members (
                        club_no, name, mobile, mobile_digits, email, rotary_id,
                        extra_json, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        item["club_no"],
                        item["name"],
                        item["mobile"],
                        item["mobile_digits"],
                        item["email"],
                        item["rotary_id"],
                        extra_json,
                        now,
                        now,
                    ),
                )
                new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                if item["rotary_id"]:
                    by_rotary[item["rotary_id"]] = new_id
                if item["mobile_digits"]:
                    by_mobile[(item["club_no"], item["mobile_digits"])] = new_id
                by_name_email[(item["club_no"], item["name"].lower(), (item["email"] or "").lower())] = new_id
                inserted += 1
                continue

            prev = conn.execute(
                """
                SELECT club_no, name, mobile, mobile_digits, email, rotary_id, extra_json
                FROM members WHERE id = ?
                """,
                (match_id,),
            ).fetchone()
            new_vals = (
                item["club_no"],
                item["name"],
                item["mobile"],
                item["mobile_digits"],
                item["email"],
                item["rotary_id"],
                extra_json,
            )
            old_vals = (
                prev["club_no"],
                prev["name"],
                prev["mobile"],
                prev["mobile_digits"],
                prev["email"],
                prev["rotary_id"],
                prev["extra_json"] or "{}",
            )
            if new_vals == old_vals:
                unchanged += 1
                continue
            conn.execute(
                """
                UPDATE members SET
                    club_no = ?, name = ?, mobile = ?, mobile_digits = ?,
                    email = ?, rotary_id = ?, extra_json = ?, updated_at = ?
                WHERE id = ?
                """,
                (*new_vals, now, match_id),
            )
            if item["rotary_id"]:
                by_rotary[item["rotary_id"]] = match_id
            if item["mobile_digits"]:
                by_mobile[(item["club_no"], item["mobile_digits"])] = match_id
            by_name_email[(item["club_no"], item["name"].lower(), (item["email"] or "").lower())] = match_id
            updated += 1

    return {
        "clubs_touched": len(seen_clubs),
        "clubs_upserted": clubs_upserted,
        "inserted": inserted,
        "updated": updated,
        "unchanged": unchanged,
        "skipped": skipped,
        "source": str(path),
    }


CLUB_NAME_ALIASES = {
    "coimbatore swasthik": "rotary club of coimbatore swasthik",
}


def _norm_club_name(name: str) -> str:
    return " ".join((name or "").strip().split()).casefold()


def _club_name_keys(name: str) -> list[str]:
    key = _norm_club_name(name)
    keys = [key]
    if key in CLUB_NAME_ALIASES:
        keys.append(CLUB_NAME_ALIASES[key])
    if key.startswith("rotary club of "):
        keys.append(key[len("rotary club of ") :])
    else:
        keys.append("rotary club of " + key)
    seen = []
    for item in keys:
        if item and item not in seen:
            seen.append(item)
    return seen


def import_club_files(path: str | Path) -> dict:
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    mapped = unmatched = skipped = updated = unchanged = 0
    now = utcnow()

    with db() as conn:
        clubs = conn.execute("SELECT club_no, club_name, pdf_url FROM clubs").fetchall()
        by_key = {}
        for club in clubs:
            for key in _club_name_keys(club["club_name"]):
                by_key.setdefault(key, club)

        pending = {}
        for row in ws.iter_rows(values_only=True):
            name = clean(row[0] if row else None)
            url = clean(row[1] if row and len(row) > 1 else None)
            if not name and not url:
                continue
            if not name or not url:
                skipped += 1
                continue
            if not url.startswith("http://") and not url.startswith("https://"):
                skipped += 1
                continue
            club = None
            for key in _club_name_keys(name):
                club = by_key.get(key)
                if club:
                    break
            if club is None:
                unmatched += 1
                continue
            pending[club["club_no"]] = (club, url)

        for club, url in pending.values():
            mapped += 1
            if (club["pdf_url"] or "") == url:
                unchanged += 1
                continue
            conn.execute(
                "UPDATE clubs SET pdf_url = ?, updated_at = ? WHERE club_no = ?",
                (url, now, club["club_no"]),
            )
            updated += 1

    return {
        "mapped": mapped,
        "updated": updated,
        "unchanged": unchanged,
        "unmatched": unmatched,
        "skipped": skipped,
        "source": str(path),
    }
